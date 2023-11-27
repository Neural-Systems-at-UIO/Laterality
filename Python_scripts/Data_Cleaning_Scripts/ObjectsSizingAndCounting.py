import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def gather_data_from_folder(folder_path, final_df, counts_df):
    """
    Populate the DataFrame template with average object sizes and counts from corresponding CSV files.
    """
    for slice_name in final_df.index:
        csv_file = f"Objects__{slice_name}.csv"
        file_path = os.path.join(folder_path, csv_file)

        if not os.path.exists(file_path):
            print(f"{csv_file} not found. Skipping...")
            continue

        try:
            df_slice = pd.read_csv(file_path, delimiter=';')
            for region in final_df.columns:
                """
                Inside the loop over regions, the next line filters the data from the slice's DataFrame
                (df_slice) to include only the rows where the "Region Name" matches the current region.
                It then selects the "Object pixels" column, which contains the size of each object/neuron in that region.
                """
                region_objects = df_slice[df_slice["Region Name"] == region]["Object pixels"]
                avg_value = region_objects.mean()
                count_value = region_objects.count()
                counts_df.at[slice_name, region] = count_value
                if not pd.isna(avg_value):
                    final_df.at[slice_name, region] = avg_value
        except Exception as e:
            print(f"Error reading {file_path}. It might be malformed. Please check the file.")
            print("Error message:", str(e))
            with open(file_path, 'r') as f:
                print("First few lines of the file:")
                for _ in range(5):
                    print(f.readline().strip())
            continue

    return final_df, counts_df

def preserve_background_colors(original_path, output_path):
    """
    Preserve the background colors from the original Excel file in the new Excel file.
    """
    # Load the original and new workbooks
    original_wb = load_workbook(original_path)
    new_wb = load_workbook(output_path)

    original_ws = original_wb.active
    new_ws = new_wb.active

    # Iterate through cells in the original worksheet and copy background colors
    for row in original_ws.iter_rows():
        for cell in row:
            if cell.fill.start_color.index in ['00FFFF00', '00008000']:
                new_ws[cell.coordinate].fill = PatternFill(start_color=cell.fill.start_color.index,
                                                           end_color=cell.fill.start_color.index,
                                                           fill_type="solid")

    # Save the changes to the new workbook
    new_wb.save(output_path)
    
def main():
    folder_path = input("Enter the path to the folder containing the CSV files: ")
    template_path = input("Please provide the path to the 'objects_adjustments_damage' XLSX file. Use this file even when analyzing colliculi objects from the same rat: ")
    template_df = pd.read_excel(template_path, engine='openpyxl', index_col=0)

    # Initialize the output DataFrames with NaN values and float data type
    final_df = pd.DataFrame(index=template_df.index, columns=template_df.columns, dtype=float)
    counts_df = pd.DataFrame(index=template_df.index, columns=template_df.columns, dtype=float)

    # DataFrames population
    final_df, counts_df = gather_data_from_folder(folder_path, final_df, counts_df)

    # Output files naming
    rat_number = input("Enter the rat number: ")
    colliculi_regions = input("Are the data related to the alignment of the colliculi? (yes/no): ")
    suffix = "_col" if colliculi_regions.lower() == "yes" else ""
    
    # Save the average object sizes to an XLSX file
    output_filename = f"{rat_number}{suffix}_objects_sizes.xlsx"
    output_file = os.path.join(folder_path, output_filename)
    final_df.to_excel(output_file, engine='openpyxl')
    
    # Save the neuron counts to a separate XLSX file
    counts_filename = f"{rat_number}{suffix}_objects_counts.xlsx"
    counts_file = os.path.join(folder_path, counts_filename)
    counts_df.to_excel(counts_file, engine='openpyxl')

    # Preserve the background colors for both files
    preserve_background_colors(template_path, output_file)
    preserve_background_colors(template_path, counts_file)
    
    print(f"Results saved to {output_file}")
    print(f"Neuron counts saved to {counts_file}")

if __name__ == "__main__":
    main()
