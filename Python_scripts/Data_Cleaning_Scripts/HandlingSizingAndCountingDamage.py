import os
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Function to get highlighted cells
def get_highlighted_cells(file_path, color):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    highlighted_cells = []

    for row in ws.iter_rows():
        for cell in row:
            if cell.fill.start_color.index == color:
                col_name = ws[f'{cell.column_letter}1'].value
                row_label = cell.row
                highlighted_cells.append((ws[f'A{row_label}'].value, col_name))

    return highlighted_cells

# Function to replace yellow highlighted cells with colliculi data
def replace_yellow_cells(original_df, colliculi_df, yellow_cells):
    for cell in yellow_cells:
        original_df.at[cell] = colliculi_df.at[cell]

# Function to process green highlighted cells
def process_green_cells(df, green_cells):
    for region in df.columns:
        green_cells_in_region = [cell for cell in green_cells if cell[1] == region]
        if not green_cells_in_region:
            continue

        # Sort the green cells by their row index
        green_cells_in_region.sort(key=lambda x: df.index.get_loc(x[0]))

        # Identify clusters of consecutive green cells
        clusters = []
        current_cluster = [green_cells_in_region[0]]
        for i in range(1, len(green_cells_in_region)):
            if df.index.get_loc(green_cells_in_region[i][0]) - df.index.get_loc(current_cluster[-1][0]) == 1:
                current_cluster.append(green_cells_in_region[i])
            else:
                clusters.append(current_cluster)
                current_cluster = [green_cells_in_region[i]]
        clusters.append(current_cluster)

        # Process each cluster
        for cluster in clusters:
            cluster_indices = [df.index.get_loc(cell[0]) for cell in cluster]
            if len(cluster) > 2:
                # Exclude values for these cells
                for cell in cluster:
                    df.at[cell] = np.nan
            else:
                # Handle one or two green cells
                # Find the previous and next non-empty cells around the cluster
                prev_idx = min(cluster_indices) - 1
                next_idx = max(cluster_indices) + 1
                prev_value = df.iloc[prev_idx][region] if prev_idx >= 0 else np.nan
                next_value = df.iloc[next_idx][region] if next_idx < len(df) else np.nan

                # Determine what value to set based on the surrounding cells
                value_to_set = np.nan
                if not np.isnan(prev_value) and not np.isnan(next_value):
                    value_to_set = (prev_value + next_value) / 2
                elif not np.isnan(prev_value):
                    value_to_set = prev_value
                elif not np.isnan(next_value):
                    value_to_set = next_value

                # Apply the determined value to all cells in the cluster
                for cell in cluster:
                    df.at[cell] = value_to_set

def preserve_background_colors(original_path, output_path):
    """
    Preserve the background colors from the original Excel file in the new Excel file.
    """
    # Load the original and new workbooks
    original_wb = load_workbook(original_path)
    new_wb = load_workbook(output_path)

    original_ws = original_wb.active
    new_ws = new_wb.active

    # Iterate through cells in the original worksheet and copy background colors
    for row in original_ws.iter_rows():
        for cell in row:
            if cell.fill.start_color.index in ['00FFFF00', '00008000']:
                new_ws[cell.coordinate].fill = PatternFill(start_color=cell.fill.start_color.index,
                                                           end_color=cell.fill.start_color.index,
                                                           fill_type="solid")

    # Save the changes to the new workbook
    new_wb.save(output_path)

# Main function to coordinate the operations
def main():
    sizes_file_path = input("Enter the path to the objects sizes table: ")
    counts_file_path = input("Enter the path to the objects counts table: ")

    # Load the data from the provided Excel files
    sizes_df = pd.read_excel(sizes_file_path, index_col=0)
    counts_df = pd.read_excel(counts_file_path, index_col=0)
    sizes_df.iloc[:, 1:] = sizes_df.iloc[:, 1:].astype(float)
    counts_df.iloc[:, 1:] = counts_df.iloc[:, 1:].astype(float)

    # Check for yellow cells
    yellow_cells_sizes = get_highlighted_cells(sizes_file_path, '00FFFF00')
    yellow_cells_counts = get_highlighted_cells(counts_file_path, '00FFFF00')

    if yellow_cells_sizes:
        colliculi_sizes_path = input("Enter the path to the colliculi sizes table: ")
        colliculi_counts_path = input("Enter the path to the colliculi counts table: ")

        colliculi_sizes_df = pd.read_excel(colliculi_sizes_path, index_col=0)
        colliculi_counts_df = pd.read_excel(colliculi_counts_path, index_col=0)

        replace_yellow_cells(sizes_df, colliculi_sizes_df, yellow_cells_sizes)
        replace_yellow_cells(counts_df, colliculi_counts_df, yellow_cells_counts)

    # Check for green cells
    green_cells_sizes = get_highlighted_cells(sizes_file_path, '00008000')
    green_cells_counts = get_highlighted_cells(counts_file_path, '00008000')
    
    process_green_cells(sizes_df, green_cells_sizes)
    process_green_cells(counts_df, green_cells_counts)

    # Determine the folder path from the original tables
    folder_path = os.path.dirname(sizes_file_path)

    # Save the corrected tables in the same path as the original tables
    rat_number = input("Enter the rat number: ")
    sizes_output_file = os.path.join(folder_path, f"{rat_number}_objects_sizes_corrected.xlsx")
    counts_output_file = os.path.join(folder_path, f"{rat_number}_objects_counts_corrected.xlsx")

    # Save the corrected DataFrames to the output files
    sizes_df.to_excel(sizes_output_file, engine='openpyxl', index=True)
    counts_df.to_excel(counts_output_file, engine='openpyxl', index=True)

    # Preserve the background colors for both files
    preserve_background_colors(sizes_file_path, sizes_output_file)
    preserve_background_colors(counts_file_path, counts_output_file)

    print(f"Corrected sizes table saved successfully as {sizes_output_file}.")
    print(f"Corrected counts table saved successfully as {counts_output_file}.")

if __name__ == "__main__":
    main()
