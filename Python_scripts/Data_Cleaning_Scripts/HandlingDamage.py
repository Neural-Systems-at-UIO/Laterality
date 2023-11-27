# Import necessary libraries
import pandas as pd
import os
import numpy as np
from openpyxl import load_workbook

# Function to retrieve cells with a specific highlight (yellow) from the Excel file
def get_highlighted_cells(file_path):
    # Load the Excel workbook [openpyxl]
    wb = load_workbook(file_path)
    # Get the active worksheet
    ws = wb.active

    highlighted_cells = []

    # Iterate over all rows in the worksheet
    for row in ws.iter_rows():
        # Iterate over each cell in the current row
        for cell in row:
            # Check if the cell has a yellow background color
            if cell.fill.start_color.index == '00FFFF00':
                # Get the column name from the first row of the current cell's column
                col_name = ws[f'{cell.column_letter}1'].value
                row_label = cell.row
                # Append the cell's row label and column name to the list
                highlighted_cells.append((ws[f'A{row_label}'].value, col_name))
    return highlighted_cells

# Function to format the slice labelling in a specific format (e.g., s001, s002, ...)
def convert_to_qc_format(slice_number):
    return f's{slice_number:03}'

# Function to check if a slice number is in the specific format mentioned above
def is_in_qc_format(slice_number):
    return True if slice_number.startswith('s') and len(slice_number) == 4 else False

# Function to map regions from the quality control DataFrame
def region_mapping(column_name, qc_extended_df):
    region_mapping = {}
    for index, row in column_name.items():
        # Check if the row value is a string
        if isinstance(row, str):
            slice_number = qc_extended_df.iloc[index, 0]
            # Split the regions separated by semicolons and strip any leading/trailing whitespace
            regions = [region.strip() for region in row.split(';')]
            region_mapping[slice_number] = regions
    return region_mapping

# Function to highlight cells based on damage and original highlighting
def highlight_cells_for_damage(row):
    colors = []
    for col in row.index:
        if (row.name, col) in modifications_damage:
            colors.append('background-color: green')
        elif (row.name, col) in original_yellow_cells:
            colors.append('background-color: yellow')
        else:
            colors.append('')
    return colors

def handle_damage(objects_df, regions_df, qc_extended_df):
    # Get the mapping of slices with damage from the quality control DataFrame
    reversed_damage_mapping = region_mapping(qc_extended_df["damage"], qc_extended_df)
    modifications_damage = {}

    # Transform reversed mapping to regions as keys for ease of processing
    damage_mapping = {}
    for slice_name, damaged_regions in reversed_damage_mapping.items():
        for region in damaged_regions:
            if region not in damage_mapping:
                damage_mapping[region] = []
            damage_mapping[region].append(slice_name)

    for region, damaged_slices in damage_mapping.items():
        # Sort slices in ascending order
        damaged_slices.sort()
        
        # Identify clusters of consecutive damaged slices
        clusters = []
        current_cluster = [damaged_slices[0]]
        for i in range(1, len(damaged_slices)):
            if objects_df.index.get_loc(damaged_slices[i]) - objects_df.index.get_loc(damaged_slices[i-1]) == 1:
                current_cluster.append(damaged_slices[i])
            else:
                clusters.append(current_cluster)
                current_cluster = [damaged_slices[i]]
        clusters.append(current_cluster)
        
        # Handle each cluster
        for cluster in clusters:
            if len(cluster) <= 2:
                # Handle edge cases
                if objects_df.index.get_loc(cluster[0]) == 0:
                    next_objects_value = objects_df.iloc[objects_df.index.get_loc(cluster[-1]) + 1][region]
                    next_regions_value = regions_df.iloc[regions_df.index.get_loc(cluster[-1]) + 1][region]
                    for slice_number in cluster:
                        objects_df.at[slice_number, region] = next_objects_value
                        regions_df.at[slice_number, region] = next_regions_value
                        modifications_damage[(slice_number, region)] = (objects_df.at[slice_number, region], next_objects_value)
                elif objects_df.index.get_loc(cluster[-1]) == len(objects_df) - 1:
                    prev_objects_value = objects_df.iloc[objects_df.index.get_loc(cluster[0]) - 1][region]
                    prev_regions_value = regions_df.iloc[regions_df.index.get_loc(cluster[0]) - 1][region]
                    for slice_number in cluster:
                        objects_df.at[slice_number, region] = prev_objects_value
                        regions_df.at[slice_number, region] = prev_regions_value
                        modifications_damage[(slice_number, region)] = (objects_df.at[slice_number, region], prev_objects_value)
                else:
                    # Compute average of neighboring undamaged slices
                    prev_objects_value = objects_df.iloc[objects_df.index.get_loc(cluster[0]) - 1][region]
                    next_objects_value = objects_df.iloc[objects_df.index.get_loc(cluster[-1]) + 1][region]
                    prev_regions_value = regions_df.iloc[regions_df.index.get_loc(cluster[0]) - 1][region]
                    next_regions_value = regions_df.iloc[regions_df.index.get_loc(cluster[-1]) + 1][region]
                    average_objects_value = (prev_objects_value + next_objects_value) / 2
                    average_regions_value = (prev_regions_value + next_regions_value) / 2
                    for slice_number in cluster:
                        objects_df.at[slice_number, region] = average_objects_value
                        regions_df.at[slice_number, region] = average_regions_value
                        modifications_damage[(slice_number, region)] = (objects_df.at[slice_number, region], average_objects_value)
            else:
                # Exclude values for these slices for the damaged region
                for slice_number in cluster:
                    objects_df.at[slice_number, region] = np.nan
                    regions_df.at[slice_number, region] = np.nan
                    modifications_damage[(slice_number, region)] = (objects_df.at[slice_number, region], np.nan)

    return objects_df, regions_df, modifications_damage

# Function to generate a filename for the modified (damage) files
def damage_filename(file_path):
    folder_path, file_name = os.path.split(file_path)
    base_name, ext = os.path.splitext(file_name)
    new_name = f"{base_name}_damage{ext}"
    return os.path.join(folder_path, new_name)

# Get file paths from the user
objects_path = input("Please enter the path to the objects file: ")
regions_path = input("Please enter the path to the regions file: ")
qc_extended_path = input("Please enter the path to the quality control (qc) file: ")

# Read the input Excel files into pandas DataFrames
objects_df = pd.read_excel(objects_path)
regions_df = pd.read_excel(regions_path)
qc_extended_df = pd.read_excel(qc_extended_path)

# Convert slice numbers to the specific format if they are not already in that format
if not objects_df['Unnamed: 0'].apply(is_in_qc_format).all():
    objects_df['Unnamed: 0'] = objects_df['Unnamed: 0'].apply(convert_to_qc_format)

if not regions_df['Unnamed: 0'].apply(is_in_qc_format).all():
    regions_df['Unnamed: 0'] = regions_df['Unnamed: 0'].apply(convert_to_qc_format)

# Set the first column as the index and convert all other columns to float type
objects_df.set_index("Unnamed: 0", inplace=True)
regions_df.set_index("Unnamed: 0", inplace=True)
objects_df.iloc[:, 1:] = objects_df.iloc[:, 1:].astype(float)
regions_df.iloc[:, 1:] = regions_df.iloc[:, 1:].astype(float)

# Get cells with the original yellow highlighting from the input Excel files
original_yellow_cells_objects = get_highlighted_cells(objects_path)
original_yellow_cells_regions = get_highlighted_cells(regions_path)
original_yellow_cells = original_yellow_cells_objects + original_yellow_cells_regions

# Handle and adjust the data for damaged regions
objects_df, regions_df, modifications_damage = handle_damage(objects_df, regions_df, qc_extended_df)

# Write the modified DataFrames to new Excel files, highlighting the modified cells
with pd.ExcelWriter(damage_filename(objects_path), engine='openpyxl') as writer:
    (objects_df.style.apply(highlight_cells_for_damage, axis=1)
     .to_excel(writer, sheet_name='Objects', index=True))

with pd.ExcelWriter(damage_filename(regions_path), engine='openpyxl') as writer:
    (regions_df.style.apply(highlight_cells_for_damage, axis=1)
     .to_excel(writer, sheet_name='Regions', index=True))

# Print a message to the user about the saved files
print(f"Dataframes saved to '{damage_filename(objects_path)}' and '{damage_filename(regions_path)}' with modified cells highlighted.")
