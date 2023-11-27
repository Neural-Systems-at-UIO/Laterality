import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

def prepare_data(df):
    df_flattened = df.melt().drop(columns='variable').reset_index()
    df_flattened.columns = ['Index', 'Count']
    return df_flattened

def plot(df_counts_flattened, df_outliers_flattened, saver_string):
    
    # Initialize the figure for subplots
    fig, axs = plt.subplots(2, 2, figsize=(15, 10))

    # 1) Plot the frequency ratio as a continuous line
    count_freq = df_counts_flattened['Count'].value_counts().sort_index()
    outlier_freq = df_outliers_flattened['Outlier_Count'].value_counts().sort_index()
    frequency_ratio = (outlier_freq / count_freq).fillna(0)
    axs[0, 0].plot(frequency_ratio.index, frequency_ratio, marker='o')
    axs[0, 0].set_title('Frequency Ratio of Outliers over Total Counts')
    axs[0, 0].set_xlabel('Object Count Value')
    axs[0, 0].set_ylabel('Frequency Ratio')
    axs[0, 0].grid(True)

    # 2) Plot the KDE
    expanded_data = []
    for count_value, ratio in frequency_ratio.items():
        expanded_data.extend([count_value] * int(round(ratio * count_freq[count_value])))
    df_expanded_data = pd.DataFrame(expanded_data, columns=['Count_Value'])
    sns.kdeplot(data=df_expanded_data, x='Count_Value', bw_adjust=0.5, ax=axs[0, 1])
    axs[0, 1].set_title('Kernel Density Estimate of Frequency Ratio Distribution')
    axs[0, 1].set_xlabel('Object Count Value')
    axs[0, 1].set_ylabel('Density')
    axs[0, 1].grid(True)

    # 3) Plot the cumulative frequency ratio with a continuous line, only for non-zero frequencies
    cumulative_count_freq = count_freq.cumsum()
    cumulative_outlier_freq = outlier_freq.cumsum()
    cumulative_outlier_freq = cumulative_outlier_freq.reindex(cumulative_count_freq.index, fill_value=0)
    cumulative_frequency_ratio = cumulative_outlier_freq / cumulative_count_freq
    non_zero_outlier_indices = cumulative_outlier_freq[cumulative_outlier_freq > 0].index
    cumulative_frequency_ratio_non_zero = cumulative_frequency_ratio[non_zero_outlier_indices]
    axs[1, 0].plot(cumulative_frequency_ratio_non_zero.index, cumulative_frequency_ratio_non_zero)
    axs[1, 0].set_title('Cumulative Ratio of Outlier Frequency over Total Count Frequency (Non-Zero Frequencies)')
    axs[1, 0].set_xlabel('Object Count Value')
    axs[1, 0].set_ylabel('Cumulative Frequency Ratio')
    axs[1, 0].grid(True)

    # 4) Replot the binned histogram with bins of size 50
    bin_size = 50
    bin_range = range(int(cumulative_frequency_ratio_non_zero.index.min()), 
                      int(cumulative_frequency_ratio_non_zero.index.max()) + bin_size, bin_size)
    binned_frequency_50 = pd.cut(cumulative_frequency_ratio_non_zero.index, bins=bin_range, right=False).value_counts().sort_index()
    binned_frequency_ratio_50 = binned_frequency_50 / binned_frequency_50.sum()
    axs[1, 1].bar(binned_frequency_ratio_50.index.astype(str), binned_frequency_ratio_50)
    axs[1, 1].set_title('Binned Histogram of Cumulative Frequency Ratio (Non-Zero Frequencies) - Bin Size 50')
    axs[1, 1].set_xlabel('Bins of Object Count Value')
    axs[1, 1].set_ylabel('Normalized Frequency Ratio')
    axs[1, 1].grid(True)

    # Adjust the layout and save the combined plot
    plt.tight_layout()
    plt.savefig(saver_string)
    plt.close()

# Ask for the rat number as user input
rat_number = input("Please enter the rat number: ")
counts_filepath = input("Please enter the counts file path: ")
dir_path = os.path.dirname(counts_filepath)

# Load the data into DataFrames and set first column as index
counts_df = pd.read_excel(counts_filepath)
counts_df_no_index = counts_df.drop(columns=['Unnamed: 0'])

# Splitting the counts DataFrame into two DataFrames
counts_df_without_totals = counts_df_no_index.iloc[:-1]  # All rows except the last
counts_df_totals = counts_df_no_index.iloc[-1:]  # Only the last row (totals)

counts_df_without_totals_flattened = prepare_data(counts_df_without_totals)
counts_df_totals_flattened = prepare_data(counts_df_totals)

# Load the workbook and the first sheet
wb = load_workbook(counts_filepath)
sheet = wb.active

# Identify all orange highlighted cells
orange_fill = 'FFFF6600'  # The color code for orange
pink_fill = 'FFFF00FF'  # The color code for pink

outliers = []
outliers_totals = []

# Iterate over rows and columns to find cells with orange fill
for row in sheet.iter_rows():
    for cell in row:
        if cell.fill.start_color.index == orange_fill:
            # Append the cell value and its index to the outliers list
            # We use the first column of the sheet as the index
            outliers.append((sheet[f'A{cell.row}'].value, cell.value))
        elif cell.fill.start_color.index == pink_fill:
            # Append the cell value and its index to the outliers list
            # We use the first column of the sheet as the index
            outliers_totals.append((sheet[f'A{cell.row}'].value, cell.value))

# Convert the outliers data into a DataFrame
df_outliers_flattened = pd.DataFrame(outliers, columns=['Index', 'Outlier_Count'])
df_outliers_totals_flattened = pd.DataFrame(outliers_totals, columns=['Index', 'Outlier_Count'])

# Define the base path for saving plots using the rat number and directory path
base_saver_string = os.path.join(dir_path, f'{rat_number}_combined_plots.png')
base_saver_string_totals = os.path.join(dir_path, f'{rat_number}_combined_plots_totals.png')

plot(counts_df_without_totals_flattened, df_outliers_flattened, base_saver_string)
plot(counts_df_totals_flattened, df_outliers_totals_flattened, base_saver_string_totals)
