import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os

def prepare_data(df):
    df_flattened = df.melt().drop(columns='variable').reset_index()
    df_flattened.columns = ['Index', 'Count']
    return df_flattened

def plot(df_counts_flattened, df_outliers_flattened, saver_string):
    
    # Initialize the figure for subplots
    fig, axs = plt.subplots(2, 2, figsize=(15, 10))

    # 1) Plot the frequency ratio as a continuous line
    count_freq = df_counts_flattened['Count'].value_counts().sort_index()
    outlier_freq = df_outliers_flattened['Outlier_Count'].value_counts().sort_index()
    frequency_ratio = (outlier_freq / count_freq).fillna(0)
    axs[0, 0].plot(frequency_ratio.index, frequency_ratio, marker='o')
    axs[0, 0].set_title('Frequency Ratio of Outliers over Total Counts')
    axs[0, 0].set_xlabel('Object Count Value')
    axs[0, 0].set_ylabel('Frequency Ratio')
    axs[0, 0].grid(True)

    # 2) Plot the KDE
    expanded_data = []
    for count_value, ratio in frequency_ratio.items():
        expanded_data.extend([count_value] * int(round(ratio * count_freq[count_value])))
    df_expanded_data = pd.DataFrame(expanded_data, columns=['Count_Value'])
    sns.kdeplot(data=df_expanded_data, x='Count_Value', bw_adjust=0.5, ax=axs[0, 1])
    axs[0, 1].set_title('Kernel Density Estimate of Frequency Ratio Distribution')
    axs[0, 1].set_xlabel('Object Count Value')
    axs[0, 1].set_ylabel('Density')
    axs[0, 1].grid(True)

    # 3) Plot the cumulative frequency ratio with a continuous line, only for non-zero frequencies
    cumulative_count_freq = count_freq.cumsum()
    cumulative_outlier_freq = outlier_freq.cumsum()
    cumulative_outlier_freq = cumulative_outlier_freq.reindex(cumulative_count_freq.index, fill_value=0)
    cumulative_frequency_ratio = cumulative_outlier_freq / cumulative_count_freq
    non_zero_outlier_indices = cumulative_outlier_freq[cumulative_outlier_freq > 0].index
    cumulative_frequency_ratio_non_zero = cumulative_frequency_ratio[non_zero_outlier_indices]
    axs[1, 0].plot(cumulative_frequency_ratio_non_zero.index, cumulative_frequency_ratio_non_zero)
    axs[1, 0].set_title('Cumulative Ratio of Outlier Frequency over Total Count Frequency (Non-Zero Frequencies)')
    axs[1, 0].set_xlabel('Object Count Value')
    axs[1, 0].set_ylabel('Cumulative Frequency Ratio')
    axs[1, 0].grid(True)

    # 4) Replot the binned histogram with bins of size 50
    bin_size = 50
    bin_range = range(int(cumulative_frequency_ratio_non_zero.index.min()), 
                      int(cumulative_frequency_ratio_non_zero.index.max()) + bin_size, bin_size)
    binned_frequency_50 = pd.cut(cumulative_frequency_ratio_non_zero.index, bins=bin_range, right=False).value_counts().sort_index()
    binned_frequency_ratio_50 = binned_frequency_50 / binned_frequency_50.sum()
    axs[1, 1].bar(binned_frequency_ratio_50.index.astype(str), binned_frequency_ratio_50)
    axs[1, 1].set_title('Binned Histogram of Cumulative Frequency Ratio (Non-Zero Frequencies) - Bin Size 50')
    axs[1, 1].set_xlabel('Bins of Object Count Value')
    axs[1, 1].set_ylabel('Normalized Frequency Ratio')
    axs[1, 1].grid(True)

    # Adjust the layout and save the combined plot
    plt.tight_layout()
    plt.savefig(saver_string)
    plt.close()

# Function to get colored cells from the workbook
def get_colored_cells(sheet, color_fill):
    cells = []
    for row in sheet.iter_rows(min_row=2):  # Skip the header
        for cell in row:
            if cell.fill.start_color.index == color_fill:
                cells.append((sheet[f'A{cell.row}'].value, cell.value))
    return cells

# Initialize comprehensive DataFrames
all_counts_without_totals = pd.DataFrame()
all_counts_totals = pd.DataFrame()
all_outliers = pd.DataFrame()
all_outliers_totals = pd.DataFrame()

# Define color codes
orange_fill = 'FFFF6600'  # The color code for orange
pink_fill = 'FFFF00FF'  # The color code for pink

# Ask for input file paths
file_paths = [input(f"Please enter the file path for rat {i}: ") for i in range(1, 5)]

# Process each file
for file_path in file_paths:
    # Load the data into DataFrame and set first column as index
    counts_df = pd.read_excel(file_path, index_col=0)
    counts_df_without_totals = counts_df.iloc[:-1]  # All rows except the last
    counts_df_totals = counts_df.iloc[-1:]  # Only the last row (totals)
    
    # Flatten the DataFrames
    counts_without_totals_flattened = prepare_data(counts_df_without_totals)
    counts_totals_flattened = prepare_data(counts_df_totals)
    
    # Append to the comprehensive DataFrames
    all_counts_without_totals = pd.concat([all_counts_without_totals, counts_without_totals_flattened])
    all_counts_totals = pd.concat([all_counts_totals, counts_totals_flattened])
    
    # Load the workbook and the first sheet
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    # Get the colored cells
    outliers = get_colored_cells(sheet, orange_fill)
    outliers_totals = get_colored_cells(sheet, pink_fill)
    
    # Convert the colored cells data into a DataFrame
    df_outliers = pd.DataFrame(outliers, columns=['Index', 'Outlier_Count'])
    df_outliers_totals = pd.DataFrame(outliers_totals, columns=['Index', 'Outlier_Count'])
    
    # Append to the comprehensive DataFrames
    all_outliers = pd.concat([all_outliers, df_outliers])
    all_outliers_totals = pd.concat([all_outliers_totals, df_outliers_totals])

folder_path = 'E:/Norvegia definitivo/v4.02_analysis'
# Define the base path for saving plots using the rat number and directory path
base_saver_string = os.path.join(folder_path, 'combined_plots.png')
base_saver_string_totals = os.path.join(folder_path, 'combined_plots_totals.png')

plot(all_counts_without_totals, all_outliers, base_saver_string)
plot(all_counts_totals, all_outliers_totals, base_saver_string_totals)
