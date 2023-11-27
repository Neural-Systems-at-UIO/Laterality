import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def calculate_totals_and_averages(sizes_path, counts_path):
    # Read the data from the Excel files
    final_df = pd.read_excel(sizes_path, index_col=0)
    counts_df = pd.read_excel(counts_path, index_col=0)

    # Check if 'Total' and 'Weighted Average' rows already exist and remove them
    if 'Total' in counts_df.index:
        counts_df = counts_df.drop('Total')
    if 'Weighted Average' in final_df.index:
        final_df = final_df.drop('Weighted Average')

    # Calculate totals for counts_df
    counts_df.loc['Total'] = counts_df.sum()

    # Calculate weighted averages for final_df
    weighted_sums = (final_df * counts_df).sum()
    total_counts = counts_df.drop('Total').sum()  # Exclude the 'Total' row itself
    weighted_averages = weighted_sums / total_counts
    final_df.loc['Weighted Average'] = weighted_averages

    return final_df, counts_df

def preserve_background_colors(original_path, output_path):
    """
    Preserve the background colors from the original Excel file in the new Excel file.
    """
    # Load the original and new workbooks
    original_wb = load_workbook(original_path)
    new_wb = load_workbook(output_path)

    original_ws = original_wb.active
    new_ws = new_wb.active

    # Iterate through cells in the original worksheet and copy background colors
    for row in original_ws.iter_rows():
        for cell in row:
            if cell.fill.start_color.index in ['00FFFF00', '00008000']:
                new_ws[cell.coordinate].fill = PatternFill(start_color=cell.fill.start_color.index,
                                                           end_color=cell.fill.start_color.index,
                                                           fill_type="solid")

    # Save the changes to the new workbook
    new_wb.save(output_path)

def main():
    # Ask user for the input files
    sizes_path = input("Enter the path to the 'objects_sizes' XLSX file: ")
    counts_path = input("Enter the path to the 'objects_counts' XLSX file: ")

    # Calculate totals and weighted averages
    final_df, counts_df = calculate_totals_and_averages(sizes_path, counts_path)

    # Save the updated data to new Excel files
    folder_path = os.path.dirname(sizes_path)
    rat_number = input("Enter the rat number: ")
    output_sizes_path = os.path.join(folder_path, f"{rat_number}_objects_sizes_corrected_with_totals.xlsx")
    output_counts_path = os.path.join(folder_path, f"{rat_number}_objects_counts_corrected_with_totals.xlsx")
    final_df.to_excel(output_sizes_path, engine='openpyxl')
    counts_df.to_excel(output_counts_path, engine='openpyxl')

    # Preserve the background colors from the original template
    preserve_background_colors(sizes_path, output_sizes_path)
    preserve_background_colors(counts_path, output_counts_path)

    print(f"Updated sizes saved to {output_sizes_path}")
    print(f"Updated counts saved to {output_counts_path}")

if __name__ == "__main__":
    main()
